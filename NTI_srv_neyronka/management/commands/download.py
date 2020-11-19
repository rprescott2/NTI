import openpyxl
from django.core.management import BaseCommand
from openpyxl import load_workbook, Workbook
from NTI_srv_neyronka.models import *


class Command(BaseCommand):
    def handle(self, *args, **options):
        ws = load_workbook('NTI_srv_neyronka/management/commands/download.xlsx')
        wb = ws['Лист1']
        wb2 = ws['Лист2']

        for row in range(2, wb.max_row):
            obj = WindTurbine(model=wb.cell(row=row, column=1).value,
                              power=float(wb.cell(row=row, column=2).value.replace('kW', '').replace(',', '.')),
                              diameter=float(wb.cell(row=row, column=3).value.replace('m', '').replace(',', '.')),
                              efficiency=float(wb.cell(row=row, column=4).value))
            obj.save()

        for row in range(2, wb2.max_row):
            obj = SolarPanel(model=wb2.cell(row=row, column=1).value,
                              power=float(wb2.cell(row=row, column=2).value.replace('W', '').replace(',', '.')),
                              efficiency=wb2.cell(row=row, column=3).value)
            obj.save()
